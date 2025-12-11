from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from app.models import (db, User, Department, Track, AcademicYear, Semester, Subject,
                        TeacherSubjectAssignment, Course, Attendance,
                        calculate_rattrapage_status, calculate_attendance_grade)
from app.utils.decorators import admin_required
from app.utils.email import send_password_creation_email
import openpyxl
from io import BytesIO

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


@admin_bp.route('/api/filter-options')
@login_required
@admin_required
def filter_options():
    """API to get options for dynamic dropdowns"""
    type = request.args.get('type')
    parent_id = request.args.get('parent_id')
    
    if not type or not parent_id:
        return jsonify([])
        
    if type == 'tracks':
        # Get tracks for a department
        tracks = Track.query.filter_by(department_id=parent_id).order_by(Track.name).all()
        return jsonify([{'id': t.id, 'name': t.name, 'level': t.level_display} for t in tracks])
        
    elif type == 'years':
        # Get academic years for a track
        years = AcademicYear.query.filter_by(track_id=parent_id).order_by(AcademicYear.order).all()
        return jsonify([{'id': y.id, 'name': y.name} for y in years])
        
    return jsonify([])


@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    """Super admin dashboard with statistics"""
    stats = {
        'departments': Department.query.count(),
        'teachers': User.query.filter_by(role='teacher').count(),
        'students': User.query.filter_by(role='student').count(),
        'tracks': Track.query.count()
    }
    
    recent_teachers = User.query.filter_by(role='teacher').order_by(User.created_at.desc()).limit(5).all()
    recent_students = User.query.filter_by(role='student').order_by(User.created_at.desc()).limit(5).all()
    departments = Department.query.all()
    
    return render_template('admin/dashboard.html', 
                          stats=stats, 
                          recent_teachers=recent_teachers,
                          recent_students=recent_students,
                          departments=departments)


# ==================== DEPARTMENT MANAGEMENT ====================

@admin_bp.route('/departments')
@login_required
@admin_required
def departments():
    """List all departments"""
    search = request.args.get('search', '').strip()
    
    query = Department.query
    if search:
        query = query.filter(Department.name.ilike(f'%{search}%'))
    
    departments = query.order_by(Department.name).all()
    return render_template('admin/departments.html', departments=departments, search=search)


@admin_bp.route('/departments/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_department():
    """Create a new department"""
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        
        if not name:
            flash('Le nom est obligatoire.', 'danger')
            return render_template('admin/department_form.html')
        
        if Department.query.filter_by(name=name).first():
            flash('Un département avec ce nom existe déjà.', 'danger')
            return render_template('admin/department_form.html')
        
        dept = Department(name=name, description=description)
        db.session.add(dept)
        db.session.commit()
        
        flash(f'Département "{name}" créé avec succès!', 'success')
        return redirect(url_for('admin.departments'))
    
    return render_template('admin/department_form.html')


@admin_bp.route('/departments/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_department(id):
    """Edit a department"""
    dept = Department.query.get_or_404(id)
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        
        if not name:
            flash('Le nom est obligatoire.', 'danger')
            return render_template('admin/department_form.html', department=dept)
        
        existing = Department.query.filter_by(name=name).first()
        if existing and existing.id != dept.id:
            flash('Un département avec ce nom existe déjà.', 'danger')
            return render_template('admin/department_form.html', department=dept)
        
        dept.name = name
        dept.description = description
        db.session.commit()
        
        flash(f'Département "{name}" modifié avec succès!', 'success')
        return redirect(url_for('admin.departments'))
    
    return render_template('admin/department_form.html', department=dept)


@admin_bp.route('/departments/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_department(id):
    """Delete a department"""
    dept = Department.query.get_or_404(id)
    name = dept.name
    
    # Check if department has teachers
    if dept.teachers:
        flash('Impossible de supprimer un département avec des enseignants assignés.', 'danger')
        return redirect(url_for('admin.departments'))
    
    db.session.delete(dept)
    db.session.commit()
    
    flash(f'Département "{name}" supprimé avec succès!', 'success')
    return redirect(url_for('admin.departments'))


# ==================== TEACHER MANAGEMENT ====================

@admin_bp.route('/teachers')
@login_required
@admin_required
def teachers():
    """List all teachers with filters"""
    department_id = request.args.get('department_id', type=int)
    search = request.args.get('search', '').strip()
    
    query = User.query.filter_by(role='teacher')
    
    if department_id:
        query = query.filter_by(department_id=department_id)
    
    if search:
        query = query.filter(
            (User.first_name.ilike(f'%{search}%')) |
            (User.last_name.ilike(f'%{search}%')) |
            (User.email.ilike(f'%{search}%')) |
            (User.matricule.ilike(f'%{search}%'))
        )
    
    teachers = query.order_by(User.last_name, User.first_name).all()
    departments = Department.query.order_by(Department.name).all()
    
    return render_template('admin/teachers.html', 
                          teachers=teachers, 
                          departments=departments,
                          selected_dept=department_id,
                          search=search)


@admin_bp.route('/teachers/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_teacher():
    """Create a new teacher"""
    departments = Department.query.order_by(Department.name).all()
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        matricule = request.form.get('matricule', '').strip()
        department_id = request.form.get('department_id', type=int)
        
        if not email or not first_name or not last_name:
            flash('Tous les champs obligatoires doivent être remplis.', 'danger')
            return render_template('admin/teacher_form.html', departments=departments)
        
        if User.query.filter_by(email=email).first():
            flash('Un utilisateur avec cet email existe déjà.', 'danger')
            return render_template('admin/teacher_form.html', departments=departments)
        
        if matricule and User.query.filter_by(matricule=matricule).first():
            flash('Un utilisateur avec ce matricule existe déjà.', 'danger')
            return render_template('admin/teacher_form.html', departments=departments)
        
        teacher = User(
            email=email,
            first_name=first_name,
            last_name=last_name,
            matricule=matricule if matricule else None,
            role='teacher',
            department_id=department_id
        )
        
        db.session.add(teacher)
        db.session.commit()
        
        # Send password creation email
        try:
            send_password_creation_email(teacher)
            db.session.commit()
            flash(f'Enseignant "{first_name} {last_name}" créé. Email envoyé!', 'success')
        except Exception as e:
            flash(f'Enseignant créé mais erreur d\'envoi email: {str(e)}', 'warning')
        
        return redirect(url_for('admin.teachers'))
    
    return render_template('admin/teacher_form.html', departments=departments)


@admin_bp.route('/teachers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_teacher(id):
    """Edit a teacher"""
    teacher = User.query.get_or_404(id)
    if teacher.role != 'teacher':
        flash('Utilisateur non trouvé.', 'danger')
        return redirect(url_for('admin.teachers'))
    
    departments = Department.query.order_by(Department.name).all()
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        matricule = request.form.get('matricule', '').strip()
        department_id = request.form.get('department_id', type=int)
        
        if not email or not first_name or not last_name:
            flash('Tous les champs obligatoires doivent être remplis.', 'danger')
            return render_template('admin/teacher_form.html', teacher=teacher, departments=departments)
        
        existing = User.query.filter_by(email=email).first()
        if existing and existing.id != teacher.id:
            flash('Un utilisateur avec cet email existe déjà.', 'danger')
            return render_template('admin/teacher_form.html', teacher=teacher, departments=departments)
        
        if matricule:
            existing_mat = User.query.filter_by(matricule=matricule).first()
            if existing_mat and existing_mat.id != teacher.id:
                flash('Un utilisateur avec ce matricule existe déjà.', 'danger')
                return render_template('admin/teacher_form.html', teacher=teacher, departments=departments)
        
        teacher.email = email
        teacher.first_name = first_name
        teacher.last_name = last_name
        teacher.matricule = matricule if matricule else None
        teacher.department_id = department_id
        db.session.commit()
        
        flash(f'Enseignant modifié avec succès!', 'success')
        return redirect(url_for('admin.teachers'))
    
    return render_template('admin/teacher_form.html', teacher=teacher, departments=departments)


@admin_bp.route('/teachers/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_teacher(id):
    """Delete a teacher"""
    teacher = User.query.get_or_404(id)
    if teacher.role != 'teacher':
        flash('Utilisateur non trouvé.', 'danger')
        return redirect(url_for('admin.teachers'))
    
    name = teacher.full_name
    
    # Remove head positions if any
    if teacher.is_dept_head and teacher.headed_department:
        teacher.headed_department.head = None
        teacher.is_dept_head = False
    
    if teacher.is_track_head and teacher.headed_track:
        teacher.headed_track.head = None
        teacher.is_track_head = False
    
    db.session.delete(teacher)
    db.session.commit()
    
    flash(f'Enseignant "{name}" supprimé avec succès!', 'success')
    return redirect(url_for('admin.teachers'))


@admin_bp.route('/teachers/import', methods=['GET', 'POST'])
@login_required
@admin_required
def import_teachers():
    """Import teachers from Excel file"""
    departments = Department.query.order_by(Department.name).all()
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Aucun fichier sélectionné.', 'danger')
            return render_template('admin/import_teachers.html', departments=departments)
        
        file = request.files['file']
        department_id = request.form.get('department_id', type=int)
        
        if file.filename == '':
            flash('Aucun fichier sélectionné.', 'danger')
            return render_template('admin/import_teachers.html', departments=departments)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Format de fichier invalide. Utilisez un fichier Excel (.xlsx).', 'danger')
            return render_template('admin/import_teachers.html', departments=departments)
        
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            
            imported = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue
                
                email = str(row[0]).strip().lower() if row[0] else None
                first_name = str(row[1]).strip() if row[1] else None
                last_name = str(row[2]).strip() if row[2] else None
                matricule = str(row[3]).strip() if len(row) > 3 and row[3] else None
                
                if not email or not first_name or not last_name:
                    errors.append(f"Ligne {row_num}: données manquantes")
                    continue
                
                if User.query.filter_by(email=email).first():
                    errors.append(f"Ligne {row_num}: email déjà existant ({email})")
                    continue

                if matricule and User.query.filter_by(matricule=matricule).first():
                    errors.append(f"Ligne {row_num}: matricule déjà existant ({matricule})")
                    continue
                
                teacher = User(
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    matricule=matricule,
                    role='teacher',
                    department_id=department_id
                )
                db.session.add(teacher)
                
                try:
                    send_password_creation_email(teacher)
                except:
                    pass
                
                imported += 1
            
            db.session.commit()
            
            if imported > 0:
                flash(f'{imported} enseignant(s) importé(s) avec succès!', 'success')
            if errors:
                flash(f'{len(errors)} erreur(s): ' + '; '.join(errors[:3]), 'warning')
            
            return redirect(url_for('admin.teachers'))
            
        except Exception as e:
            flash(f'Erreur lors de l\'import: {str(e)}', 'danger')
            return render_template('admin/import_teachers.html', departments=departments)
    
    return render_template('admin/import_teachers.html', departments=departments)


# ==================== DEPARTMENT HEAD MANAGEMENT ====================

@admin_bp.route('/departments/<int:id>/head', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_department_head(id):
    """Assign or change department head"""
    dept = Department.query.get_or_404(id)
    teachers = User.query.filter_by(role='teacher', department_id=id).order_by(User.last_name).all()
    
    if request.method == 'POST':
        teacher_id = request.form.get('teacher_id', type=int)
        
        if teacher_id:
            new_head = User.query.get(teacher_id)
            if not new_head or new_head.department_id != dept.id:
                flash('Enseignant invalide.', 'danger')
                return redirect(url_for('admin.assign_department_head', id=id))
            
            # Remove old head
            old_head = dept.head
            if old_head:
                old_head.is_dept_head = False
                old_head.headed_department_id = None
            
            # Assign new head
            new_head.is_dept_head = True
            new_head.headed_department_id = dept.id
            
            db.session.commit()
            flash(f'{new_head.full_name} est maintenant chef du département "{dept.name}".', 'success')
        else:
            # Remove current head
            if dept.head:
                dept.head.is_dept_head = False
                dept.head.headed_department_id = None
                db.session.commit()
                flash('Chef de département retiré.', 'info')
        
        return redirect(url_for('admin.departments'))
    
    return render_template('admin/assign_head.html', department=dept, teachers=teachers, type='department')


# ==================== STUDENTS VIEW ====================

@admin_bp.route('/students')
@login_required
@admin_required
def students():
    """List all students with filters"""
    track_id = request.args.get('track_id', type=int)
    search = request.args.get('search', '').strip()
    
    query = User.query.filter_by(role='student')
    
    if track_id:
        query = query.filter(User.enrolled_tracks.any(id=track_id))
    
    if search:
        query = query.filter(
            (User.first_name.ilike(f'%{search}%')) |
            (User.last_name.ilike(f'%{search}%')) |
            (User.email.ilike(f'%{search}%'))
        )
    
    students = query.order_by(User.last_name, User.first_name).all()
    tracks = Track.query.order_by(Track.name).all()
    
    return render_template('admin/students.html', 
                          students=students, 
                          tracks=tracks,
                          selected_track=track_id,
                          search=search)


# ==================== TRACK (FILIÈRE) MANAGEMENT ====================

@admin_bp.route('/tracks')
@login_required
@admin_required
def tracks():
    """List all tracks"""
    department_id = request.args.get('department_id', type=int)
    search = request.args.get('search', '').strip()
    
    query = Track.query
    if department_id:
        query = query.filter_by(department_id=department_id)
    if search:
        query = query.filter(Track.name.ilike(f'%{search}%'))
    
    tracks = query.order_by(Track.name).all()
    departments = Department.query.order_by(Department.name).all()
    
    return render_template('admin/tracks.html', 
                          tracks=tracks, 
                          departments=departments,
                          selected_dept=department_id,
                          search=search)


@admin_bp.route('/tracks/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_track():
    """Create a new track"""
    departments = Department.query.order_by(Department.name).all()
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        level = request.form.get('level', 'licence')
        description = request.form.get('description', '').strip()
        department_id = request.form.get('department_id', type=int)
        auto_generate = request.form.get('auto_generate') == 'on'
        
        if not name or not department_id:
            flash('Tous les champs obligatoires doivent être remplis.', 'danger')
            return render_template('admin/track_form.html', departments=departments)
        
        if Track.query.filter_by(name=name, department_id=department_id).first():
            flash('Une filière avec ce nom existe déjà dans ce département.', 'danger')
            return render_template('admin/track_form.html', departments=departments)
        
        track = Track(name=name, level=level, description=description, department_id=department_id)
        db.session.add(track)
        db.session.commit()
        
        # Auto-generate academic structure if requested
        if auto_generate:
            structure = Track.get_academic_structure(level)
            for i, (code, year_name) in enumerate(structure['years'], 1):
                year = AcademicYear(name=year_name, order=i, track_id=track.id)
                db.session.add(year)
                db.session.commit()
                
                for j in range(1, structure['semesters_per_year'] + 1):
                    sem_num = (i - 1) * structure['semesters_per_year'] + j
                    semester = Semester(name=f'Semestre {sem_num}', order=j, academic_year_id=year.id)
                    db.session.add(semester)
            db.session.commit()
        
        flash(f'Filière "{name}" créée avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/track_form.html', departments=departments)


@admin_bp.route('/tracks/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_track(id):
    """Edit a track"""
    track = Track.query.get_or_404(id)
    departments = Department.query.order_by(Department.name).all()
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        level = request.form.get('level', 'licence')
        description = request.form.get('description', '').strip()
        department_id = request.form.get('department_id', type=int)
        
        if not name:
            flash('Le nom est obligatoire.', 'danger')
            return render_template('admin/track_form.html', track=track, departments=departments)
        
        existing = Track.query.filter_by(name=name, department_id=department_id).first()
        if existing and existing.id != track.id:
            flash('Une filière avec ce nom existe déjà.', 'danger')
            return render_template('admin/track_form.html', track=track, departments=departments)
        
        track.name = name
        track.level = level
        track.description = description
        track.department_id = department_id
        db.session.commit()
        
        flash('Filière modifiée avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/track_form.html', track=track, departments=departments)


@admin_bp.route('/tracks/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_track(id):
    """Delete a track"""
    track = Track.query.get_or_404(id)
    name = track.name
    
    # Remove track head role if any
    if track.head:
        track.head.is_track_head = False
        track.head.headed_track_id = None
    
    db.session.delete(track)
    db.session.commit()
    
    flash(f'Filière "{name}" supprimée avec succès!', 'success')
    return redirect(url_for('admin.tracks'))


@admin_bp.route('/tracks/<int:id>/head', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_track_head(id):
    """Assign track head"""
    track = Track.query.get_or_404(id)
    teachers = User.query.filter_by(role='teacher', department_id=track.department_id).order_by(User.last_name).all()
    
    if request.method == 'POST':
        teacher_id = request.form.get('teacher_id', type=int)
        
        if teacher_id:
            new_head = User.query.get(teacher_id)
            if not new_head:
                flash('Enseignant invalide.', 'danger')
                return redirect(url_for('admin.assign_track_head', id=id))
            
            # Remove old head
            if track.head:
                track.head.is_track_head = False
                track.head.headed_track_id = None
            
            # Assign new head
            new_head.is_track_head = True
            new_head.headed_track_id = track.id
            
            db.session.commit()
            flash(f'{new_head.full_name} est maintenant chef de la filière "{track.name}".', 'success')
        else:
            # Remove current head
            if track.head:
                track.head.is_track_head = False
                track.head.headed_track_id = None
                db.session.commit()
                flash('Chef de filière retiré.', 'info')
        
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/assign_track_head.html', track=track, teachers=teachers)


@admin_bp.route('/tracks/<int:id>/teachers', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_track_teachers(id):
    """Assign teachers to a track"""
    track = Track.query.get_or_404(id)
    teachers = User.query.filter_by(role='teacher', department_id=track.department_id).order_by(User.last_name).all()
    
    if request.method == 'POST':
        selected_ids = request.form.getlist('teacher_ids', type=int)
        
        # Clear current assignments and add selected teachers
        track.assigned_teachers = []
        for tid in selected_ids:
            teacher = User.query.get(tid)
            if teacher:
                track.assigned_teachers.append(teacher)
        
        db.session.commit()
        flash('Enseignants assignés avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/assign_track_teachers.html', track=track, teachers=teachers)


# ==================== ACADEMIC STRUCTURE MANAGEMENT ====================

# Track structure route removed - merged into tracks list


@admin_bp.route('/tracks/<int:track_id>/year/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_academic_year(track_id):
    """Create academic year"""
    track = Track.query.get_or_404(track_id)
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        order = request.form.get('order', type=int, default=1)
        
        if not name:
            flash('Le nom est obligatoire.', 'danger')
            return render_template('admin/academic_year_form.html', track=track)
        
        year = AcademicYear(name=name, order=order, track_id=track_id)
        db.session.add(year)
        db.session.commit()
        
        flash(f'Année "{name}" créée avec succès!', 'success')
        flash(f'Année "{name}" créée avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/academic_year_form.html', track=track)


@admin_bp.route('/year/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_academic_year(id):
    """Edit academic year"""
    year = AcademicYear.query.get_or_404(id)
    track = year.track
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        order = request.form.get('order', type=int, default=1)
        
        if not name:
            flash('Le nom est obligatoire.', 'danger')
            return render_template('admin/academic_year_form.html', year=year, track=track)
        
        year.name = name
        year.order = order
        db.session.commit()
        
        flash('Année modifiée avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/academic_year_form.html', year=year, track=track)


@admin_bp.route('/year/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_academic_year(id):
    """Delete academic year"""
    year = AcademicYear.query.get_or_404(id)
    track_id = year.track_id
    name = year.name
    
    db.session.delete(year)
    db.session.commit()
    
    flash(f'Année "{name}" supprimée avec succès!', 'success')
    return redirect(url_for('admin.tracks'))


@admin_bp.route('/year/<int:year_id>/semester/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_semester(year_id):
    """Create semester"""
    year = AcademicYear.query.get_or_404(year_id)
    track = year.track
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        order = request.form.get('order', type=int, default=1)
        
        if not name:
            flash('Le nom est obligatoire.', 'danger')
            return render_template('admin/semester_form.html', year=year, track=track)
        
        semester = Semester(name=name, order=order, academic_year_id=year_id)
        db.session.add(semester)
        db.session.commit()
        
        flash(f'Semestre "{name}" créé avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/semester_form.html', year=year, track=track)


@admin_bp.route('/semester/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_semester(id):
    """Delete semester"""
    semester = Semester.query.get_or_404(id)
    track_id = semester.academic_year.track_id
    name = semester.name
    
    db.session.delete(semester)
    db.session.commit()
    
    flash(f'Semestre "{name}" supprimé avec succès!', 'success')
    return redirect(url_for('admin.tracks'))


@admin_bp.route('/semester/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_semester(id):
    """Edit semester"""
    semester = Semester.query.get_or_404(id)
    year = semester.academic_year
    track = year.track
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        order = request.form.get('order', type=int, default=1)
        
        if not name:
            flash('Le nom est obligatoire.', 'danger')
            return render_template('admin/semester_form.html', semester=semester, year=year, track=track)
        
        semester.name = name
        semester.order = order
        db.session.commit()
        
        flash('Semestre modifié avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/semester_form.html', semester=semester, year=year, track=track)


# ==================== SUBJECT MANAGEMENT ====================

@admin_bp.route('/semester/<int:semester_id>/subject/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_subject(semester_id):
    """Create subject"""
    semester = Semester.query.get_or_404(semester_id)
    track = semester.academic_year.track
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        code = request.form.get('code', '').strip().upper()
        description = request.form.get('description', '').strip()
        total_cm = request.form.get('total_cm', type=int, default=0)
        total_td = request.form.get('total_td', type=int, default=0)
        total_tp = request.form.get('total_tp', type=int, default=0)
        
        if not name:
            flash('Le nom est obligatoire.', 'danger')
            return render_template('admin/subject_form.html', semester=semester, track=track)
        
        # Auto-generate code if empty
        if not code:
            # 1. Track Abbreviation (First 4 chars of valid word, or INFO default)
            track_words = track.name.upper().split()
            abbrev = track_words[0][:4] if track_words else 'GENI'
            for word in track_words:
                if word not in ['GENIE', 'MASTER', 'LICENCE', 'DOCTORAT']:
                    abbrev = word[:4]
                    break
            
            # 2. Level (L1, L2, M1...) from Academic Year name
            # Assuming simplified year names or fallback logic
            level_initial = track.level[0].upper() if track.level else 'L'
            year_order = semester.academic_year.order
            level_code = f"{level_initial}{year_order}"
            
            # 3. Number (Count + 1)
            # Count subjects in the same academic year to maintain sequence
            existing_count = 0
            for sem in semester.academic_year.semesters:
                existing_count += len(sem.subjects)
            number = existing_count + 1
            
            code = f"{abbrev}-{level_code}-{number:02d}"

        subject = Subject(
            name=name, code=code, description=description,
            semester_id=semester_id,
            total_cm=total_cm, total_td=total_td, total_tp=total_tp
        )
        db.session.add(subject)
        db.session.commit()
        
        flash(f'Matière "{name}" créée avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    # Pre-generate code for display in form
    pre_filled_code = ''
    try:
        # 1. Track Abbreviation
        track_words = track.name.upper().split()
        abbrev = track_words[0][:4] if track_words else 'GENI'
        for word in track_words:
            if word not in ['GENIE', 'MASTER', 'LICENCE', 'DOCTORAT']:
                abbrev = word[:4]
                break
        
        # 2. Level Code
        level_initial = track.level[0].upper() if track.level else 'L'
        year_order = semester.academic_year.order
        level_code = f"{level_initial}{year_order}"
        
        # 3. Number
        existing_count = 0
        for sem in semester.academic_year.semesters:
            existing_count += len(sem.subjects)
        number = existing_count + 1
        
        pre_filled_code = f"{abbrev}-{level_code}-{number:02d}"
    except Exception as e:
        print(f"Error generating code: {e}")

    return render_template('admin/subject_form.html', semester=semester, track=track, pre_filled_code=pre_filled_code)


@admin_bp.route('/subject/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_subject(id):
    """Edit subject"""
    subject = Subject.query.get_or_404(id)
    semester = subject.semester
    track = semester.academic_year.track
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        code = request.form.get('code', '').strip().upper()
        description = request.form.get('description', '').strip()
        total_cm = request.form.get('total_cm', type=int, default=0)
        total_td = request.form.get('total_td', type=int, default=0)
        total_tp = request.form.get('total_tp', type=int, default=0)
        
        if not name or not code:
            flash('Le nom et le code sont obligatoires.', 'danger')
            return render_template('admin/subject_form.html', subject=subject, semester=semester, track=track)
        
        subject.name = name
        subject.code = code
        subject.description = description
        subject.total_cm = total_cm
        subject.total_td = total_td
        subject.total_tp = total_tp
        db.session.commit()
        
        flash('Matière modifiée avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/subject_form.html', subject=subject, semester=semester, track=track)


@admin_bp.route('/subject/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_subject(id):
    """Delete subject"""
    subject = Subject.query.get_or_404(id)
    track_id = subject.semester.academic_year.track_id
    name = subject.name
    
    db.session.delete(subject)
    db.session.commit()
    
    flash(f'Matière "{name}" supprimée avec succès!', 'success')
    return redirect(url_for('admin.tracks'))


@admin_bp.route('/subject/<int:id>/assign', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_subject_teacher(id):
    """Assign teacher to subject"""
    subject = Subject.query.get_or_404(id)
    track = subject.semester.academic_year.track
    teachers = User.query.filter_by(role='teacher', department_id=track.department_id).order_by(User.last_name).all()
    
    if request.method == 'POST':
        teacher_id = request.form.get('teacher_id', type=int)
        teaches_cm = 'teaches_cm' in request.form
        teaches_td = 'teaches_td' in request.form
        teaches_tp = 'teaches_tp' in request.form
        
        if not teacher_id:
            flash('Sélectionnez un enseignant.', 'danger')
            return render_template('admin/assign_subject_teacher.html', subject=subject, teachers=teachers, track=track)
        
        # Check existing assignment
        assignment = TeacherSubjectAssignment.query.filter_by(
            teacher_id=teacher_id, subject_id=id
        ).first()
        
        if assignment:
            assignment.teaches_cm = teaches_cm
            assignment.teaches_td = teaches_td
            assignment.teaches_tp = teaches_tp
        else:
            assignment = TeacherSubjectAssignment(
                teacher_id=teacher_id, subject_id=id,
                teaches_cm=teaches_cm, teaches_td=teaches_td, teaches_tp=teaches_tp
            )
            db.session.add(assignment)
        
        db.session.commit()
        flash('Enseignant assigné avec succès!', 'success')
        return redirect(url_for('admin.tracks'))
    
    return render_template('admin/assign_subject_teacher.html', subject=subject, teachers=teachers, track=track)


# ==================== GLOBAL STATISTICS ====================

@admin_bp.route('/statistics')
@login_required
@admin_required
def global_statistics():
    """Global establishment statistics"""
    # Get all tracks with their statistics
    tracks_data = []
    for track in Track.query.all():
        track_stats = {
            'track': track,
            'students_count': len(track.students),
            'teachers_count': len(track.assigned_teachers),
            'subjects_count': 0,
            'courses_count': 0,
            'attendance_rate': 0
        }
        
        total_present = 0
        total_attendance = 0
        
        for year in track.academic_years:
            for semester in year.semesters:
                track_stats['subjects_count'] += len(semester.subjects)
                for subject in semester.subjects:
                    courses = Course.query.filter_by(subject_id=subject.id, status='completed').all()
                    track_stats['courses_count'] += len(courses)
                    for course in courses:
                        for att in course.attendances:
                            total_attendance += 1
                            if att.status == 'present':
                                total_present += 1
        
        if total_attendance > 0:
            track_stats['attendance_rate'] = round(total_present / total_attendance * 100, 1)
        
        tracks_data.append(track_stats)
    
    # Overall stats
    overall = {
        'departments': Department.query.count(),
        'tracks': Track.query.count(),
        'teachers': User.query.filter_by(role='teacher').count(),
        'students': User.query.filter_by(role='student').count(),
        'subjects': Subject.query.count(),
        'courses_completed': Course.query.filter_by(status='completed').count()
    }
    
    return render_template('admin/statistics.html', tracks_data=tracks_data, overall=overall)
# ==================== STUDENT MANAGEMENT ====================

@admin_bp.route('/students/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_student():
    """Create a new student"""
    departments = Department.query.order_by(Department.name).all()
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        matricule = request.form.get('matricule', '').strip()
        track_id = request.form.get('track_id', type=int)
        academic_year_id = request.form.get('academic_year_id', type=int)
        
        if not email or not first_name or not last_name:
            flash('Tous les champs obligatoires doivent être remplis.', 'danger')
            return render_template('admin/student_form.html', departments=departments)
        
        if User.query.filter_by(email=email).first():
            flash('Un utilisateur avec cet email existe déjà.', 'danger')
            return render_template('admin/student_form.html', departments=departments)
        
        if matricule and User.query.filter_by(matricule=matricule).first():
            flash('Un utilisateur avec ce matricule existe déjà.', 'danger')
            return render_template('admin/student_form.html', departments=departments)
        
        student = User(
            email=email,
            first_name=first_name,
            last_name=last_name,
            matricule=matricule if matricule else None,
            role='student',
            current_year_id=academic_year_id if academic_year_id else None
        )
        
        # Assign to track if selected
        if track_id:
            track = Track.query.get(track_id)
            if track:
                student.enrolled_tracks.append(track)
                student.department_id = track.department_id # Assign department for reference
        
        db.session.add(student)
        db.session.commit()
        
        # Send password creation email
        try:
            send_password_creation_email(student)
            flash(f'Étudiant "{first_name} {last_name}" créé. Email envoyé!', 'success')
        except Exception as e:
            flash(f'Étudiant créé mais erreur d\'envoi email: {str(e)}', 'warning')
        
        return redirect(url_for('admin.students'))
    
    return render_template('admin/student_form.html', departments=departments)


@admin_bp.route('/students/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_student(id):
    """Edit a student"""
    student = User.query.get_or_404(id)
    if student.role != 'student':
        flash('Utilisateur non trouvé.', 'danger')
        return redirect(url_for('admin.students'))
    
    departments = Department.query.order_by(Department.name).all()
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        matricule = request.form.get('matricule', '').strip()
        track_id = request.form.get('track_id', type=int)
        academic_year_id = request.form.get('academic_year_id', type=int)
        
        if not email or not first_name or not last_name:
            flash('Tous les champs obligatoires doivent être remplis.', 'danger')
            return render_template('admin/student_form.html', student=student, departments=departments)
        
        existing = User.query.filter_by(email=email).first()
        if existing and existing.id != student.id:
            flash('Un utilisateur avec cet email existe déjà.', 'danger')
            return render_template('admin/student_form.html', student=student, departments=departments)
        
        if matricule:
            existing_mat = User.query.filter_by(matricule=matricule).first()
            if existing_mat and existing_mat.id != student.id:
                flash('Un utilisateur avec ce matricule existe déjà.', 'danger')
                return render_template('admin/student_form.html', student=student, departments=departments)
        
        student.email = email
        student.first_name = first_name
        student.last_name = last_name
        student.matricule = matricule if matricule else None
        student.current_year_id = academic_year_id if academic_year_id else None
        
        # Update track enrollment
        student.enrolled_tracks = [] # Clear existing
        if track_id:
            track = Track.query.get(track_id)
            if track:
                student.enrolled_tracks.append(track)
                student.department_id = track.department_id
                
        db.session.commit()
        
        flash(f'Étudiant modifié avec succès!', 'success')
        return redirect(url_for('admin.students'))
    
    return render_template('admin/student_form.html', student=student, departments=departments)


@admin_bp.route('/students/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_student(id):
    """Delete a student"""
    student = User.query.get_or_404(id)
    if student.role != 'student':
        flash('Utilisateur non trouvé.', 'danger')
        return redirect(url_for('admin.students'))
    
    name = student.full_name
    db.session.delete(student)
    db.session.commit()
    
    flash(f'Étudiant "{name}" supprimé avec succès!', 'success')
    return redirect(url_for('admin.students'))


@admin_bp.route('/students/import', methods=['GET', 'POST'])
@login_required
@admin_required
def import_students():
    """Import students from Excel"""
    departments = Department.query.order_by(Department.name).all()
    
    if request.method == 'POST':
        track_id = request.form.get('track_id', type=int)
        academic_year_id = request.form.get('academic_year_id', type=int)

        if not track_id or not academic_year_id:
            flash('Veuillez sélectionner une filière et une année.', 'danger')
            return render_template('admin/import_students.html', departments=departments)
            
        track = Track.query.get(track_id)
        if not track:
             flash('Filière invalide.', 'danger')
             return render_template('admin/import_students.html', departments=departments)

        if 'file' not in request.files:
            flash('Aucun fichier sélectionné.', 'danger')
            return render_template('admin/import_students.html', departments=departments)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('Aucun fichier sélectionné.', 'danger')
            return render_template('admin/import_students.html', departments=departments)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Format de fichier invalide. Utilisez un fichier Excel (.xlsx).', 'danger')
            return render_template('admin/import_students.html', departments=departments)
        
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            
            imported = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip if email is missing
                    continue
                
                email = str(row[0]).strip().lower() if row[0] else None
                first_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
                last_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
                matricule = str(row[3]).strip() if len(row) > 3 and row[3] else None
                
                if not email or not first_name or not last_name:
                    errors.append(f"Ligne {row_num}: données manquantes (Email, Prénom, Nom requis)")
                    continue
                
                if User.query.filter_by(email=email).first():
                    errors.append(f"Ligne {row_num}: email déjà existant ({email})")
                    continue

                if matricule and User.query.filter_by(matricule=matricule).first():
                    errors.append(f"Ligne {row_num}: matricule déjà existant ({matricule})")
                    continue
                
                student = User(
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    matricule=matricule,
                    role='student',
                    current_year_id=academic_year_id,
                    department_id=track.department_id
                )
                
                # Assign to selected track
                student.enrolled_tracks.append(track)
                
                db.session.add(student)
                
                try:
                    send_password_creation_email(student)
                except:
                    pass
                
                imported += 1
            
            db.session.commit()
            
            if imported > 0:
                flash(f'{imported} étudiant(s) importé(s) dans {track.name} (Année ID: {academic_year_id}) !', 'success')
            if errors:
                flash(f'{len(errors)} erreur(s): ' + '; '.join(errors[:3]), 'warning')
            
            return redirect(url_for('admin.students'))
            
        except Exception as e:
            flash(f'Erreur lors de l\'import: {str(e)}', 'danger')
            return render_template('admin/import_students.html', departments=departments)
    
    return render_template('admin/import_students.html', departments=departments)
